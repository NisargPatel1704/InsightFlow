import io
from datetime import date
from flask import Blueprint, send_file, request, current_app
from flask_login import login_required
from app.models import Sale, Customer

reports_bp = Blueprint("reports", __name__, template_folder="../../templates")


@reports_bp.route("/sales/excel")
@login_required
def sales_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    status = request.args.get("status", "all")
    query = Sale.query.order_by(Sale.sale_date.desc())
    if status != "all":
        query = query.filter(Sale.status == status)
    sales = query.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    headers = ["Invoice #", "Date", "Customer", "Payment Method", "Status", "Subtotal", "Tax", "Total"]
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, sale in enumerate(sales, start=2):
        ws.cell(row=row_idx, column=1, value=sale.invoice_number)
        ws.cell(row=row_idx, column=2, value=sale.sale_date.strftime("%Y-%m-%d"))
        ws.cell(row=row_idx, column=3, value=sale.customer.name if sale.customer else "—")
        ws.cell(row=row_idx, column=4, value=sale.payment_method or "—")
        ws.cell(row=row_idx, column=5, value=sale.status.title())
        ws.cell(row=row_idx, column=6, value=round(sale.subtotal, 2))
        ws.cell(row=row_idx, column=7, value=round(sale.tax, 2))
        ws.cell(row=row_idx, column=8, value=round(sale.total, 2))

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    total_row = len(sales) + 2
    ws.cell(row=total_row, column=7, value="Total:").font = Font(bold=True)
    ws.cell(row=total_row, column=8, value=round(sum(s.total for s in sales), 2)).font = Font(bold=True)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"insightflow_sales_report_{date.today().isoformat()}.xlsx"
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@reports_bp.route("/sales/pdf")
@login_required
def sales_pdf():
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    status = request.args.get("status", "all")
    query = Sale.query.order_by(Sale.sale_date.desc())
    if status != "all":
        query = query.filter(Sale.status == status)
    sales = query.all()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.6 * inch, bottomMargin=0.6 * inch)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("TitleStyle", parent=styles["Title"], textColor=colors.HexColor("#4F46E5"))

    elements = [
        Paragraph(f"{current_app.config.get('APP_NAME')} — Sales Report", title_style),
        Paragraph(f"Generated on {date.today().strftime('%B %d, %Y')}", styles["Normal"]),
        Spacer(1, 0.25 * inch),
    ]

    data = [["Invoice #", "Date", "Customer", "Status", "Total"]]
    for sale in sales:
        data.append([
            sale.invoice_number,
            sale.sale_date.strftime("%Y-%m-%d"),
            (sale.customer.name if sale.customer else "—")[:28],
            sale.status.title(),
            f"${sale.total:,.2f}",
        ])

    total_revenue = sum(s.total for s in sales)
    data.append(["", "", "", "Total", f"${total_revenue:,.2f}"])

    table = Table(data, colWidths=[1.1 * inch, 1 * inch, 2.3 * inch, 1 * inch, 1.1 * inch])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F46E5")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#F3F4F6")]),
        ("LINEBELOW", (0, 0), (-1, 0), 1, colors.HexColor("#4F46E5")),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("LINEABOVE", (0, -1), (-1, -1), 1, colors.HexColor("#4F46E5")),
        ("ALIGN", (-1, 0), (-1, -1), "RIGHT"),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)

    filename = f"insightflow_sales_report_{date.today().isoformat()}.pdf"
    return send_file(buffer, as_attachment=True, download_name=filename, mimetype="application/pdf")


@reports_bp.route("/invoice/<int:sale_id>/pdf")
@login_required
def invoice_pdf(sale_id):
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    sale = Sale.query.get_or_404(sale_id)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.7 * inch, bottomMargin=0.7 * inch)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("TitleStyle", parent=styles["Title"], textColor=colors.HexColor("#4F46E5"))
    label_style = ParagraphStyle("LabelStyle", parent=styles["Normal"], textColor=colors.HexColor("#6B7280"))

    elements = [
        Paragraph(current_app.config.get("APP_NAME"), title_style),
        Paragraph(f"Invoice {sale.invoice_number}", styles["Heading2"]),
        Spacer(1, 0.15 * inch),
        Paragraph(f"Bill to: {sale.customer.name if sale.customer else '—'}", styles["Normal"]),
        Paragraph(f"Date: {sale.sale_date.strftime('%B %d, %Y')}", label_style),
        Paragraph(f"Status: {sale.status.title()}", label_style),
        Spacer(1, 0.3 * inch),
    ]

    data = [["Product", "Qty", "Unit Price", "Line Total"]]
    for item in sale.items:
        data.append([
            item.product.name if item.product else "—",
            str(item.quantity),
            f"${item.unit_price:,.2f}",
            f"${item.line_total:,.2f}",
        ])
    data.append(["", "", "Subtotal", f"${sale.subtotal:,.2f}"])
    data.append(["", "", "Tax", f"${sale.tax:,.2f}"])
    data.append(["", "", "Total", f"${sale.total:,.2f}"])

    table = Table(data, colWidths=[2.6 * inch, 0.8 * inch, 1.3 * inch, 1.3 * inch])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F46E5")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("FONTNAME", (2, -3), (-1, -1), "Helvetica-Bold"),
        ("LINEABOVE", (0, -1), (-1, -1), 1, colors.HexColor("#4F46E5")),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"invoice_{sale.invoice_number}.pdf",
        mimetype="application/pdf",
    )
